from openpyxl import load_workbook
from pprint import pprint
from Google import make_service

"""
create event. 
"""


def event(name, dates, times0, times1, event_description, location, calendar_id):
    services = make_service()
    if (str(times1) == str(None)) or (str(times0) == str(None)):
        event_request_body = {
            'summary': name,
            'location': location,
            'description': event_description,
            'start': {
                'date': dates,
                'timeZone': 'Asia/Jerusalem',
            },
            'end': {
                'date': dates,
                'timeZone': 'Asia/Jerusalem',
            },
            'colorId': 5,
            'status': 'confirmed',
            'transparency': 'opaque',
            'visibility': 'private',
            'location': '',
            'attachments': [
                {
                    'fileUrl': '  ',
                    'title': ''
                }
            ],
            'reminders': {
                'useDefault': False,
                'overrides': [
                    {'method': 'popup', 'minutes': 60 * 24},
                ],
            },
        }
    else:
        event_request_body = {
            'summary': name,
            'location': location,
            'description': event_description,
            'start': {
                'dateTime': str(dates.date()) + "T" + str(times0) + ":00-" + str(times1),
                'timeZone': 'Asia/Jerusalem',
            },
            'end': {
                'dateTime': str(dates.date()) + "T" + str(times0) + ":00-" + str(times1),
                'timeZone': 'Asia/Jerusalem',
            },
            'colorId': 5,
            'status': 'confirmed',
            'transparency': 'opaque',
            'visibility': 'private',
            'location': '',
            'attachments': [
                {
                    'fileUrl': '  ',
                    'title': ''
                }
            ],
            'reminders': {
                'useDefault': False,
                'overrides': [
                    {'method': 'popup', 'minutes': 60 * 24},
                ],
            },
        }

    response = services.events().insert(
        calendarId=calendar_id,
        body=event_request_body
    ).execute()


def send_to_google_calendar(file_path, calendar_id):
    wb = load_workbook(filename=file_path)
    for sheet in wb.worksheets:
        if sheet.title == "source":
            continue

        row_count = sheet.max_row
        for row in range(2, row_count - 1):
            print(sheet.cell(row=row, column=3).value)
            b_dates = sheet.cell(row=row, column=3).value
            times0 = sheet.cell(row=row, column=4).value
            times1 = sheet.cell(row=row, column=5).value
            event_name = sheet.cell(row=row, column=6).value
            event_description = sheet.cell(row=row, column=7).value
            location = sheet.cell(row=row, column=8).value
            phone_number = sheet.cell(row=row, column=9).value
            pprint("/////////////////////////////////////////////////////////")
            pprint(event_description)
            pprint("/////////////////////////////////////////////////////////")
            event(name=event_name, dates=b_dates, times0=times0, times1=times1,
                  event_description=event_description, location=location, calendar_id=calendar_id)
