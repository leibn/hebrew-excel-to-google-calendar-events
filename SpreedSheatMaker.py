from openpyxl import load_workbook
import os.path
from pyluach import dates
from pyluach.dates import HebrewDate
from pprint import pprint
"""inside project imports"""
import gematria


"dictionary to converts the hebrew date"
STANDARD_ENCODING = {
    "תשרי": 7,
    "חשוון": 8,
    "כסלו": 9,
    "טבת": 10,
    "שבט": 11,
    "אדר": 12,
    "אדרב": 13,
    "ניסן": 1,
    "אייר": 2,
    "סיוון": 3,
    "תמוז": 4,
    "אב": 5,
    "אלול": 6
}
inv_STANDARD_ENCODING = {v: k for k, v in STANDARD_ENCODING.items()}


def decode_word(word: str) -> int:
    word = word.replace(u'\xa0', u'')
    if word.strip() in STANDARD_ENCODING:
        return STANDARD_ENCODING[word]
    else:
        print("the month : " + str(word) + "is not valid month"
                                           "enter alternate month number:")
        pprint(STANDARD_ENCODING)
        alternate_month = int(input())
        print(decode_num(alternate_month) + "הוזן בהצלחה")
        return alternate_month


def decode_num(user_num):
    if int(user_num) < 14:
        return inv_STANDARD_ENCODING[user_num]
    else:
        print("חודש יכול להיות מספר בין 1 ל13 לפי הטבלה אנא נסה להזין בשנית את המספר של החודש הנכון")
        print("enter alternate month number:")
        pprint(STANDARD_ENCODING)
        alternate_month = int(input())
        print(decode_num(alternate_month))
        return alternate_month


def generate_dates(n_years, month, day):
    hebrew_today_year = dates.HebrewDate.today().year - 1
    try:
        hebrew_n_birth_date = HebrewDate(year=hebrew_today_year + n_years, month=month, day=day)
    except ValueError:
        try:
            hebrew_n_birth_date = HebrewDate(year=hebrew_today_year + n_years, month=month, day=day - 1)
        except ValueError:
            hebrew_n_birth_date = HebrewDate(year=hebrew_today_year + n_years, month=month - 1, day=day)

    greg_n_birth_date = hebrew_n_birth_date.to_greg()
    return greg_n_birth_date


def make_work_book(raw_date_path, next_n_years):
    wb = load_workbook(filename=raw_date_path)
    sheets_list = []
    version = 1
    dest_filename = "more" + str(next_n_years) + 'years_of_hebrew_dates_v' + str(version) + '.xlsx'
    indicator = os.path.exists(dest_filename)
    while indicator:
        version += 1
        dest_filename = "more" + str(next_n_years) + 'years_of_hebrew_dates_v' + str(version) + '.xlsx'
        indicator = os.path.exists(dest_filename)
    wb.active
    sheet = wb.worksheets[0]
    sheet.title = "source"
    row_count = sheet.max_row
    column_index = 1
    hebrew_today = dates.HebrewDate.today()

    for n in range(1, next_n_years + 1):
        first_row = ('index', 'pearson name', 'תאריך שהיום העברי יוצא לפי הלוח הגאורגיאני',
                     ' 	שעת התחלה (אם אירוע של יום שלם לא להזין שעות)',
                     'שעת סיום (אם אירוע של יום שלם לא להזין שעות)'
                     , 'שם אירוע', 'תיאור האירוע', '	כתובת מגורים', 'מספר פלאפון', '"יום" תאריך הולדת עברי',
                     '	"חודש" תאריך הולדת עברי)', 'עברי במספרים')

        # create the sheet in workbook(excel file)
        sheets_list.append(wb.create_sheet(title=str(
            dates.HebrewDate(year=hebrew_today.year - 1 + n, month=hebrew_today.month, day=hebrew_today.day).year)))
        # copy the value to columns header
        sheets_list[n - 1].append(first_row)
        for row in range(2, row_count - 1):
            # assignments of the data to sheets
            sheets_list[n - 1].cell(row=row, column=column_index).value = row
            sheets_list[n - 1].cell(row=row, column=column_index + 1).value = \
                sheet.cell(row=row, column=3).value
            sheets_list[n - 1].cell(row=row, column=column_index + 5).value = str(sheet['c' + str(row)].value) +" HebrewBirthDay "
            sheets_list[n - 1].cell(row=row, column=column_index + 7).value = "שורה 7"
            sheets_list[n - 1].cell(row=row, column=column_index + 8).value = "שורה 8"
            sheets_list[n - 1].cell(row=row, column=column_index + 9).value = \
                sheet.cell(column=1, row=row).value
            sheets_list[n - 1].cell(row=row, column=column_index + 10).value = \
                sheet.cell(column=2, row=row).value

            month = decode_word(str(sheet.cell(column=2, row=row).value))
            day = gematria.Gematria.decode_word(sheet.cell(column=1, row=row).value)

            sheets_list[n - 1].cell(row=row, column=column_index + 11).value = str(month) + "/" + str(day)
            sheets_list[n - 1].cell(row=row, column=column_index + 2).value = \
                str(generate_dates(n_years=n, month=month, day=day))
            sheets_list[n - 1].cell(row=row, column=column_index + 6).value = \
                "יום הולדת עברי של " + sheet['c' + str(row)].value + "\n\t" + \
                "כתובת: " + sheets_list[n - 1].cell(row=row, column=column_index + 7).value + "\n\t" + \
                "טלפון: " + sheets_list[n - 1].cell(row=row, column=column_index + 8).value + "\n\t" + \
                "תאריך עברי: " + sheets_list[n - 1].cell(row=row, column=column_index + 9).value + "' ב" + \
                sheets_list[n - 1].cell(row=row, column=column_index + 10).value + "\n\t"

    wb.save(filename=dest_filename)
    wb.close()

    return dest_filename
